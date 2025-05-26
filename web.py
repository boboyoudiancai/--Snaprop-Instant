from flask import Flask, render_template, request, url_for, jsonify, send_file, session, redirect

from datetime import datetime
import os
from pathlib import Path
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from report.ocr import OCR_Table
from config.path_config import UPLOAD_FOLDER, OCR_PATH, REPORT_PATH
from estimator import Estimator
from record.record import Record
from llm.llm_manager import QianwenManager
import time
from werkzeug.security import generate_password_hash, check_password_hash
import secrets

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'templates')
app = Flask(__name__, template_folder=TEMPLATE_DIR)
app.secret_key = secrets.token_hex(16)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OCR_PATH'] = OCR_PATH
app.config['REPORT_PATH'] = REPORT_PATH

user_sessions = {}  # 全局用户会话存储

# 模拟用户数据库（实际应使用真实数据库）
mock_users_db = {
    "testuser": {
        "password_hash": generate_password_hash("password123"),
        "uid": "1001"
    }
}


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # 验证用户信息
        user = mock_users_db.get(username)
        if user and check_password_hash(user['password_hash'], password):
            session.clear()  # 清除旧会话
            session['uid'] = user['uid']
            session['username'] = username
            session.permanent = True  # 设置持久化
            session.modified = True  # 强制会话写入
            get_or_create_user_session()
            return redirect(url_for('index'))
        else:
            error = '用户名或密码错误，请重新输入！'

    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()  # 清除会话
    return redirect(url_for('login'))


@app.route('/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')

        if not username or not password:
            return jsonify({
                "success": False,
                "error": "缺少必要参数"
            }), 400

        if username in mock_users_db:
            return jsonify({
                "success": False,
                "error": "用户已存在"
            }), 409

        # 创建新用户
        mock_users_db[username] = {
            "password_hash": generate_password_hash(password),
            "uid": str(len(mock_users_db) + 1001)
        }

        return jsonify({
            "success": True,
            "message": "注册成功"
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


def login_required(f):
    def wrapper(*args, **kwargs):
        print(f"Checking session in login_required: {session}")
        print(f"Request path: {request.path}")
        print(f"Request method: {request.method}")
        print(f"Request headers: {dict(request.headers)}")

        if 'username' not in session:
            print("No username in session, redirecting to login")
            # 检查是否是 AJAX 请求
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': '用户未登录'}), 401
            return redirect(url_for('login'))

        print(f"User {session['username']} is authenticated")
        return f(*args, **kwargs)

    wrapper.__name__ = f.__name__
    return wrapper


def get_or_create_user_session():
    """获取或创建用户会话"""
    uid = session.get('uid')
    print(f"Getting user session for uid: {uid}")

    if not uid:
        raise ValueError("UID not found in session")  # 防止意外情况

    if uid not in user_sessions:
        # 为用户创建独立实例
        user_sessions[uid] = {
            'record': Record(uid),
            'llm': QianwenManager(),
            'estimator': Estimator(uid, None, None)
        }
        user_sessions[uid]['estimator'] = Estimator(
            uid,
            user_sessions[uid]['record'],
            user_sessions[uid]['llm']
        )

    return user_sessions[uid]


def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify(success=False, error="No file part")

    file = request.files['file']
    image_type = request.form.get('image_type', 'property_photo')

    if file.filename == '':
        return jsonify(success=False, error="No selected file")

    if file and allowed_file(file.filename):
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        # 使用pathlib处理路径
        save_path = Path(app.config['UPLOAD_FOLDER']) / filename
        save_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            # 保存文件
            with open(save_path, 'wb') as f:
                f.write(file.read())
            # 验证
            if not save_path.exists():
                return jsonify(success=False, error="文件保存失败"), 500
        except Exception as e:
            print(f"保存异常: {str(e)}")
            return jsonify(success=False, error=str(e)), 500
        web_url = url_for('static', filename=f'uploads/{filename}')

        if image_type == 'property_photo':
            user_sessions[session['uid']]['record'].add_field(f"{UPLOAD_FOLDER}/{filename}")
        elif image_type == 'property_cert':
            user_sessions[session['uid']]['record'].add_property(f"{UPLOAD_FOLDER}/{filename}")

        return jsonify(success=True, url=web_url, filename=filename)

    return jsonify(success=False, error="Invalid file type")


@app.route('/')
@login_required
def index():
    return render_template('home.html')


@app.route('/app/chat', methods=['POST'])
def chat():
    try:
        # print("chat")
        data = request.get_json()
        if not data or 'message' not in data:
            return jsonify({"error": "Invalid request format"}), 400

        # 设置用户消息到estimator
        user_sessions[session['uid']]['estimator'].set_user_message(data['message'])
        # 执行交互分析
        user_sessions[session['uid']]['estimator'].interact_estimator()
        # 获取处理结果
        response = user_sessions[session['uid']]['estimator'].get_analyst_result()[0]
        print(response)
        if user_sessions[session['uid']]['estimator'].is_report():
            return jsonify({
                "response": response,
                "isReport": True
            })
        else:
            return jsonify({
                "response": response,
                "isReport": False
            })
    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 500


@app.route('/app/get_ocr_result', methods=['POST'])
def get_ocr_result():
    try:
        # print("get_ocr_result")
        data = request.get_json()
        if not data or 'filename' not in data:
            return jsonify({"error": "Invalid request format"}), 400
        filename = data['filename']
        filepath = Path(app.config['OCR_PATH']) / filename
        print(filepath)
        user_sessions[session['uid']]['estimator'].interact_table(filepath)
        response = user_sessions[session['uid']]['estimator'].get_analyst_result()[0]
        print(response)
        return jsonify({
            "response": response
        })
    except Exception as e:
        print("error", e)
        return jsonify({"error": str(e)}), 500


@app.route('/download/<filename>')
def download_file(filename):
    try:
        print(filename)
        ext = Path(filename).suffix.lower()
        if ext == ".pdf":
            file_path = Path(app.config['REPORT_PATH']) / filename
        elif ext == ".xlsx":
            file_path = Path(app.config['OCR_PATH']) / filename
        else:
            return jsonify(success=False, error="错误的文件名"), 404
        if not file_path.exists():
            return jsonify(success=False, error="文件不存在"), 404
        return send_file(
            str(file_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


# 自动调整行高
def auto_adjust_row_height(ws):
    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value:
                # 计算换行数
                lines = str(cell.value).count('\n') + 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row[0].row].height = 15 * max_lines


@app.route('/save_ocr_data', methods=['POST'])
def save_ocr_data():
    try:
        data = request.json
        final_data = data.get('ocr_data', {})
        print(final_data)

        if not final_data:
            return jsonify(success=False, error="空数据"), 400

        # 创建唯一文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ocr_data_{timestamp}.xlsx"
        file_path = Path(app.config['OCR_PATH']) / filename

        # 确保目录存在
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # 直接写入本地文件
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, sheet_data in final_data.items():
                # 保存数据
                df = pd.DataFrame(
                    sheet_data['data'],
                    columns=sheet_data.get('columns', [])
                )
                df.to_excel(
                    writer,
                    sheet_name=sheet_name[:31],
                    index=False,
                    header=False
                )

                # 恢复合并单元格
                worksheet = writer.sheets[sheet_name[:31]]
                # 设置全局居中样式
                alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

                # 遍历所有单元格（从第1行开始）
                for row in worksheet.iter_rows(min_row=1):  # min_row=1跳过标题行
                    for cell in row:
                        cell.alignment = alignment
                for merge in sheet_data.get('merges', []):
                    worksheet.merge_cells(
                        start_row=merge['start_row'] + 1,  # 转换回1-based索引
                        end_row=merge['end_row'] + 1,
                        start_column=merge['start_col'] + 1,
                        end_column=merge['end_col'] + 1
                    )
                    cell = worksheet.cell(
                        row=merge['start_row'] + 1,
                        column=merge['start_col'] + 1
                    )
                    cell.alignment = alignment
                auto_adjust_row_height(worksheet)  # 调用行高调整

        file_stat = file_path.stat()
        target_path = file_path.resolve()  # 转换为绝对路径
        max_retries = 10  # 添加最大重试次数

        # 调试输出实际检测路径
        print(f"正在检测文件路径: {target_path}")

        for _ in range(max_retries):
            if target_path.exists():
                print("文件验证成功")
                break
            print("文件尚未就绪，等待...")
            time.sleep(1)
        else:
            return jsonify(success=False, error="文件生成超时"), 500

        user_sessions[session['uid']]['record'].add_property_ocr(str(target_path))  # 统一使用绝对路径
        return jsonify(
            success=True,
            file_info={
                "name": filename,
                "size": f"{file_stat.st_size / 1024:.1f}KB",
                "time": datetime.fromtimestamp(file_stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "url": url_for('download_file', filename=filename)
            }
        )
        # return send_file(
        #     str(file_path),
        #     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        #     as_attachment=True,
        #     download_name=filename
        # )
    except Exception as e:
        # 打印异常信息，方便排查问题
        print(f"保存Excel文件失败: {str(e)}")
        return jsonify(success=False, error=str(e)), 500


@app.route('/get_report', methods=['POST'])
def get_report():
    try:
        filename = user_sessions[session['uid']]['estimator'].get_report()[0]
        file_path = Path(app.config['REPORT_PATH']) / filename
        file_stat = file_path.stat()
        return jsonify(
            success=True,
            file_info={
                "name": filename,
                "size": f"{file_stat.st_size / 1024:.1f}KB",
                "time": datetime.fromtimestamp(file_stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "url": url_for('download_file', filename=filename)
            }
        )

    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@app.route('/upload/ocr', methods=['POST'])
def ocr():
    data = request.get_json()
    if not data or 'ocr_img' not in data:
        return jsonify({"error": "Invalid request format"}), 400
    img_name = data['ocr_img']
    print(img_name)
    try:
        ocr_save_path = OCR_Table().trans_to_xlsx(img_name)
        print(ocr_save_path)
        # ocr_save_path = "static/ocr_tables/20250220123822_cropped_image_OCR.xlsx"#one sheet
        # ocr_save_path = "static/ocr_tables/20250220113936_cropped_image_20250220113938.xlsx"  # two sheets
        ocr_data = OCR_Table().trans_to_dict(ocr_save_path)
        print(ocr_data)
        return jsonify({
            "success": True,
            "data": ocr_data
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/clean_records', methods=['POST'])
def clean_records():
    try:
        user_sessions[session['uid']]['record'].clear()
        return jsonify({
            "success": True,
            "message": f"成功清理"
        })

    except Exception as e:
        print(f"清理失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"清理失败: {str(e)}"
        }), 500


if __name__ == '__main__':
    app.run(debug=True)
