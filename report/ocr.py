import datetime
import io
import os
import sys
import time
from typing import List

import openpyxl
import requests
from alibabacloud_ocr_api20210707.client import Client as ocr_api20210707Client
from alibabacloud_tea_openapi import models as open_api_models
from alibabacloud_ocr_api20210707 import models as ocr_api_20210707_models
from alibabacloud_tea_util import models as util_models
from alibabacloud_tea_util.client import Client as UtilClient
from PIL import Image
from pathlib import Path
import pandas as pd
from config.ocr_config import ocr_api_id, ocr_api_secret
from config.path_config import OCR_PATH,UPLOAD_FOLDER


class OCR_Table:
    def __init__(self, api_id: str = ocr_api_id, api_secret: str = ocr_api_secret):
        """初始化OCR处理器
        """
        self.api_id = api_id
        self.api_secret = api_secret
        # self.temp_dir = Path(OCR_PATH)
        # self.temp_dir.mkdir(exist_ok=True)

    def create_client(self) -> ocr_api20210707Client:
        """
        使用AK&SK初始化账号Client
        @return: Client
        @throws Exception
        """
        # 工程代码泄露可能会导致 AccessKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考。
        # 建议使用更安全的 STS 方式，更多鉴权访问方式请参见：https://help.aliyun.com/document_detail/378659.html。
        config = open_api_models.Config(
            # 必填，请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_ID。,
            access_key_id=self.api_id,
            # 必填，请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_SECRET。,
            access_key_secret=self.api_secret
        )
        # Endpoint 请参考 https://api.aliyun.com/product/ocr-api
        config.endpoint = f'ocr-api.cn-hangzhou.aliyuncs.com'
        return ocr_api20210707Client(config)

    def trans_to_str(self, img_path) -> str | None:
        with open(img_path, 'rb') as img_file:
            binary_data = img_file.read()
        # img=Image.open(img_path)
        img_bytes = io.BytesIO(binary_data)
        file_name, file_ext = os.path.splitext(img_path)
        # img.save(img_bytes,format=file_ext[1:])
        # img_bytes.seek(0)
        client = self.create_client()
        recognize_all_text_request = ocr_api_20210707_models.RecognizeAllTextRequest(
            body=img_bytes,
            type='Table',
            table_config=ocr_api_20210707_models.RecognizeAllTextRequestTableConfig(output_table_excel=True)
        )
        runtime = util_models.RuntimeOptions()
        try:
            resp = client.recognize_all_text_with_options(recognize_all_text_request, runtime)
            # ConsoleClient.log(UtilClient.to_jsonstring(resp))
            return UtilClient.to_jsonstring(resp)
        except Exception as error:
            # 此处仅做打印展示，请谨慎对待异常处理，在工程项目中切勿直接忽略异常。
            # 错误 message
            print(error.message)
            # 诊断地址
            print(error.data.get("Recommend"))
            UtilClient.assert_as_string(error.message)

    def trans_to_url(self, img_path) -> str | None:
        url = eval(self.trans_to_str(img_path))['body']["Data"]["SubImages"][0]["TableInfo"]["TableExcel"]  # important
        return url

    def trans_to_path(self, img_path):
        filename = f"{Path(img_path).stem}_OCR.xlsx"
        filepath = f"{OCR_PATH}/{filename}"
        return filepath

    def trans_to_xlsx(self, img_name):
        img_path = f"{UPLOAD_FOLDER}/{img_name}"
        print(img_path)
        url = self.trans_to_url(img_path)
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        save_path = self.trans_to_path(img_name)
        with open(save_path, 'wb') as f:
            f.write(response.content)

        return save_path

    # def trans_to_dict(self, save_path) -> dict:
    #     wb = openpyxl.load_workbook(save_path)
    #     result = {}
    #
    #     for sheet_name in wb.sheetnames:
    #         df = pd.DataFrame(wb[sheet_name].values)
    #         result[sheet_name] = {
    #             "columns": df.columns.tolist(),
    #             "data": df.values.tolist()
    #         }
    #     return result

    def trans_to_dict(self, save_path) -> dict:
        wb = openpyxl.load_workbook(save_path)
        result = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            df = pd.DataFrame(ws.values)

            # 获取合并单元格信息
            merges = []
            for merge_range in ws.merged_cells.ranges:
                merges.append({
                    'start_row': merge_range.min_row - 1,  # 转换为0-based索引
                    'end_row': merge_range.max_row - 1,
                    'start_col': merge_range.min_col - 1,
                    'end_col': merge_range.max_col - 1
                })

            result[sheet_name] = {
                "columns": df.columns.tolist(),
                "data": df.values.tolist(),
                "merges": merges  # 新增合并信息
            }
        return result

    def trans_to_df(self, save_path) -> list | None:
        wb = openpyxl.load_workbook(save_path)
        # for sheet_name in wb.sheetnames:
        #     sheet = wb[sheet_name]
        #     for merged_cell_range in list(sheet.merged_cells.ranges):
        #         merged_cell = merged_cell_range.start_cell
        #         sheet.unmerge_cells(range_string=merged_cell_range.coord)
        #         for row_index, col_index in merged_cell_range.cells:
        #             cell = sheet.cell(row=row_index, column=col_index)
        #             cell.value = merged_cell.value
        # wb.save(file_path)
        # wb.close()
        # xlsx = pd.ExcelFile(file_path)
        data = []
        for sheet_name in wb.sheetnames:
            data.append(list(wb[sheet_name].values))
        return data


if (__name__ == '__main__'):
    img_path = "20250306173756_cropped_image.png"
    # save_path = "../static/ocr_tables/ocr_data_20250226_221523.xlsx"
    ocr_processor = OCR_Table()  # 正确实例化
    # print(ocr_processor.trans_to_df(save_path))  # 通过实例调用
    print(ocr_processor.trans_to_xlsx(img_path))
