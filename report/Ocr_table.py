import datetime
import io
import os
import sys
import time
from typing import List

import openpyxl
import requests
from alibabacloud_ocr_api20210707.client import Client as ocr_api20210707Client
from alibabacloud_tea_openapi import models as open_api_models
from alibabacloud_ocr_api20210707 import models as ocr_api_20210707_models
from alibabacloud_tea_util import models as util_models
# from alibabacloud_tea_console.client import Client as ConsoleClient
from alibabacloud_tea_util.client import Client as UtilClient
from PIL import Image
import pandas as pd
from config.ocr_config import ocr_api_id, ocr_api_secret

class OCR_Table:
    def __init__(self):
        pass
    @staticmethod
    def create_client(ak_id, ak_secret) -> ocr_api20210707Client:
        """
        使用AK&SK初始化账号Client
        @return: Client
        @throws Exception
        """
        # 工程代码泄露可能会导致 AccessKey 泄露，并威胁账号下所有资源的安全性。以下代码示例仅供参考。
        # 建议使用更安全的 STS 方式，更多鉴权访问方式请参见：https://help.aliyun.com/document_detail/378659.html。
        config = open_api_models.Config(
            # 必填，请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_ID。,
            access_key_id=ak_id,
            # 必填，请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_SECRET。,
            access_key_secret=ak_secret
        )
        # Endpoint 请参考 https://api.aliyun.com/product/ocr-api
        config.endpoint = f'ocr-api.cn-hangzhou.aliyuncs.com'
        return ocr_api20210707Client(config)
    @staticmethod
    def trans_to_str(ak_id,ak_secret,img_path) -> str | None:
        with open(img_path, 'rb') as img_file:
            binary_data = img_file.read()
        # img=Image.open(img_path)
        img_bytes=io.BytesIO(binary_data)
        file_name, file_ext = os.path.splitext(img_path)
        # img.save(img_bytes,format=file_ext[1:])
        # img_bytes.seek(0)
        client = OCR_Table.create_client(ak_id,ak_secret)
        recognize_all_text_request = ocr_api_20210707_models.RecognizeAllTextRequest(
            body=img_bytes,
            type='Table',
            table_config=ocr_api_20210707_models.RecognizeAllTextRequestTableConfig(output_table_excel=True)
        )
        runtime = util_models.RuntimeOptions()
        try:
            resp = client.recognize_all_text_with_options(recognize_all_text_request, runtime)
            # ConsoleClient.log(UtilClient.to_jsonstring(resp))
            return UtilClient.to_jsonstring(resp)
        except Exception as error:
            # 此处仅做打印展示，请谨慎对待异常处理，在工程项目中切勿直接忽略异常。
            # 错误 message
            print(error.message)
            # 诊断地址
            print(error.data.get("Recommend"))
            UtilClient.assert_as_string(error.message)
    @staticmethod
    def trans_to_url(ak_id,ak_secret,img_path) -> str | None:
        url = eval(OCR_Table.trans_to_str(ak_id,ak_secret,img_path))['body']["Data"]["SubImages"][0]["TableInfo"]["TableExcel"]  # important
        return url
    @staticmethod
    def trans_to_xlsx(ak_id,ak_secret,img_path) -> str | None:
        response = requests.get(OCR_Table.trans_to_url(ak_id,ak_secret,img_path))
        file_path = time.strftime("%Y%m%d%H%M%S", time.localtime()) + '.xlsx'
        with open(file_path, 'wb') as file:
            file.write(response.content)
            return file_path
    @staticmethod
    def trans_to_df(ak_id, ak_secret, img_path) -> list | None:
        file_path=OCR_Table.trans_to_xlsx(ak_id,ak_secret,img_path)
        wb = openpyxl.load_workbook(file_path)
        # for sheet_name in wb.sheetnames:
        #     sheet = wb[sheet_name]
        #     for merged_cell_range in list(sheet.merged_cells.ranges):
        #         merged_cell = merged_cell_range.start_cell
        #         sheet.unmerge_cells(range_string=merged_cell_range.coord)
        #         for row_index, col_index in merged_cell_range.cells:
        #             cell = sheet.cell(row=row_index, column=col_index)
        #             cell.value = merged_cell.value
        # wb.save(file_path)
        # wb.close()
        # xlsx = pd.ExcelFile(file_path)
        data = []
        for sheet_name in wb.sheetnames:
            data.append(list(wb[sheet_name].values))
        return data


    # @staticmethod
    # async def main_async(self,
    #     args: List[str],
    # ) -> str|None:
    #     ak_id = 0
    #     ak_secret = 1
    #     img_path = 2
    #     img = Image.open(args[img_path])
    #     img_bytes = io.BytesIO()
    #     file_name, file_ext = os.path.splitext(args[img_path])
    #     img.save(img_bytes, format=file_ext[1:])
    #     img_bytes.seek(0)
    #     client = OCR.create_client(args[ak_id],args[ak_secret])
    #     recognize_all_text_request = ocr_api_20210707_models.RecognizeAllTextRequest(
    #         body=img_bytes,
    #         type='Table',
    #         table_config=ocr_api_20210707_models.RecognizeAllTextRequestTableConfig(output_table_excel=True)
    #     )
    #     runtime = util_models.RuntimeOptions()
    #     try:
    #         resp = await client.recognize_all_text_with_options_async(recognize_all_text_request, runtime)
    #         ConsoleClient.log(UtilClient.to_jsonstring(resp))
    #         return UtilClient.to_jsonstring(resp)
    #     except Exception as error:
    #         # 此处仅做打印展示，请谨慎对待异常处理，在工程项目中切勿直接忽略异常。
    #         # 错误 message
    #         print(error.message)
    #         # 诊断地址
    #         print(error.data.get("Recommend"))
    #         UtilClient.assert_as_string(error.message)

if __name__ == '__main__':
    img_path="D:/sitp_work/web/static/uploads/20250306174923_cropped_image.png"
    ocr_processor = OCR_Table()  # 正确实例化
    # print(ocr_processor.trans_to_df(save_path))  # 通过实例调用
    print(ocr_processor.trans_to_xlsx(ocr_api_id,ocr_api_secret,img_path))
